from openpyxl import load_workbook

def cargar_comentarios(path_datos):
    """
    Lee la hoja 'Comentarios' del archivo de datos y devuelve un diccionario
    indexado por factura.
    """

    wb = load_workbook(path_datos, data_only=True)
    hoja = wb['Comentarios']

    comentarios = {}

    # Recorremos desde la fila 2 (omitimos encabezados)
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        factura = fila[0]
        ruta = fila[1]
        comentario = fila[3]

        # Saltar filas sin número de factura
        if not factura:
            continue

        comentarios[str(factura).strip()] = {
            'ruta': int(ruta) if ruta not in (None, '') else None,
            'comentario': comentario or ''
        }

    return comentarios


def cargar_info_cliente(path_datos):
    """
    Lee la hoja 'Datos' del archivo de datos y devuelve un diccionario
    con la información general del cliente y sus rutas por departamento.
    """

    wb = load_workbook(path_datos, data_only=True)
    hoja = wb['Datos']

    clientes = {}

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        codigo = fila[0]

        # Saltar filas sin código de cliente
        if not codigo:
            continue

        codigo = str(codigo).strip()

        clientes[codigo] = {
            # Información general del cliente
            'negocio': fila[2] or '',
            'tipo_pago': fila[3] or '',

            # Información por departamento
            'Abacer': {
                'ruta': fila[5] or '',
                'dia': fila[6] or ''
            },
            'Brigar': {
                'ruta': fila[7] or '',
                'dia': fila[8] or ''
            },
            'Marlboro': {
                'ruta': fila[9] or '',
                'dia': fila[10] or ''
            },
            'Holanda': {
                'ruta': fila[11] or '',
                'dia': fila[12] or ''
            },
            'Piso': {
                'ruta': fila[13] or '',
                'dia': fila[14] or ''
            }
        }

    return clientes
