from openpyxl import load_workbook
from openpyxl.styles import numbers
from datetime import datetime
import os
import shutil

# =========================
# Orden fijo de rutas
# =========================
ORDEN_RUTAS = [703, 701, 705, 706, 702, 709, 710, 711, 712, 708, 714]


def exportar_excel(registros, plantilla_path, salida_path):
    """
    Genera el archivo final de antigüedad usando una plantilla base.
    """

    # =========================
    # Cargar plantilla
    # =========================
    wb = load_workbook(plantilla_path)
    hoja = wb['Antigüedad']

    fila_inicio = 10

    # =========================
    # Resumen general
    # =========================
    resumen = calcular_resumen(registros)

    # =========================
    # Llenado hoja general
    # =========================
    for idx, r in enumerate(registros):
        fila = fila_inicio + idx

        hoja.cell(row=fila, column=1, value=r['factura'])
        hoja.cell(row=fila, column=2, value=r['departamento'])
        hoja.cell(row=fila, column=3, value=r['id_cliente'])
        hoja.cell(row=fila, column=4, value=r['cliente'])
        hoja.cell(row=fila, column=5, value=r['negocio'])

        celda_fecha = hoja.cell(row=fila, column=6, value=r['fecha'])
        celda_fecha.number_format = 'DD/MM/YYYY'

        hoja.cell(row=fila, column=7, value=r['antiguedad'])

        celda_total = hoja.cell(row=fila, column=8, value=r['total'])
        celda_total.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        hoja.cell(row=fila, column=9, value=r['tipo_pago'])
        hoja.cell(row=fila, column=10, value=r['ruta'])
        hoja.cell(row=fila, column=11, value=r['dia'])
        hoja.cell(row=fila, column=13, value=r['comentarios'])

    # =========================
    # Escribir resumen general
    # =========================
    filas_resumen = {
        'total': 10,
        'menor_7': 11,
        'mayor_7': 12,
        'mayor_14': 13,
    }

    for clave, fila in filas_resumen.items():
        hoja.cell(row=fila, column=16, value=resumen[clave]['facturas'])
        celda_saldo = hoja.cell(row=fila, column=17, value=resumen[clave]['saldo'])
        celda_saldo.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # =========================
    # Procesar hojas por ruta
    # =========================
    rutas_dict = agrupar_por_ruta(registros)

    for ruta in ORDEN_RUTAS:
        if ruta not in rutas_dict:
            continue

        if str(ruta) not in wb.sheetnames:
            continue

        hoja_ruta = wb[str(ruta)]
        registros_ruta = rutas_dict[ruta]

        escribir_hoja_ruta(hoja_ruta, registros_ruta)

        resumen_ruta = calcular_resumen(registros_ruta)
        escribir_resumen_ruta(hoja_ruta, resumen_ruta)

    # =========================
    # Limpiar hojas auxiliares
    # =========================
    for nombre in ('Datos', 'Comentarios'):
        if nombre in wb.sheetnames:
            wb.remove(wb[nombre])

    # =========================
    # Guardar archivo final
    # =========================
    wb.save(salida_path)


# ==================================================
# Cálculo de resumen de facturas
# ==================================================
def calcular_resumen(registros):
    resumen = {
        'total': {'facturas': 0, 'saldo': 0},
        'menor_7': {'facturas': 0, 'saldo': 0},
        'mayor_7': {'facturas': 0, 'saldo': 0},
        'mayor_14': {'facturas': 0, 'saldo': 0},
    }

    for r in registros:
        antiguedad = r['antiguedad']
        total = r['total'] or 0

        resumen['total']['facturas'] += 1
        resumen['total']['saldo'] += total

        if antiguedad < 7:
            clave = 'menor_7'
        elif antiguedad < 14:
            clave = 'mayor_7'
        else:
            clave = 'mayor_14'

        resumen[clave]['facturas'] += 1
        resumen[clave]['saldo'] += total

    return resumen


# ==================================================
# Agrupación de registros por ruta
# ==================================================
def agrupar_por_ruta(registros):
    rutas = {}

    for r in registros:
        try:
            ruta = int(r['ruta'])
        except (TypeError, ValueError):
            continue

        rutas.setdefault(ruta, []).append(r)

    return rutas


# ==================================================
# Escritura de hoja individual por ruta
# ==================================================
def escribir_hoja_ruta(hoja, registros):
    fila_inicio = 5

    for idx, r in enumerate(registros):
        fila = fila_inicio + idx

        hoja.cell(row=fila, column=1, value=r['factura'])
        hoja.cell(row=fila, column=2, value=r['cliente'])
        hoja.cell(row=fila, column=3, value=r['negocio'])

        celda_fecha = hoja.cell(row=fila, column=4, value=r['fecha'])
        celda_fecha.number_format = 'DD/MM/YYYY'

        hoja.cell(row=fila, column=5, value=r['antiguedad'])

        celda_total = hoja.cell(row=fila, column=6, value=r['total'])
        celda_total.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        hoja.cell(row=fila, column=7, value=r['tipo_pago'])
        hoja.cell(row=fila, column=8, value=r['dia'])
        hoja.cell(row=fila, column=9, value=r['fisico'])
        hoja.cell(row=fila, column=10, value=r['comentarios'])


# ==================================================
# Escritura de resumen en hoja de ruta
# ==================================================
def escribir_resumen_ruta(hoja, resumen):
    filas = {
        'total': 5,
        'menor_7': 6,
        'mayor_7': 7,
        'mayor_14': 8,
    }

    for clave, fila in filas.items():
        hoja.cell(row=fila, column=13, value=resumen[clave]['facturas'])
        celda_saldo = hoja.cell(row=fila, column=14, value=resumen[clave]['saldo'])
        celda_saldo.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


# ==================================================
# Generación de archivos individuales por ruta
# ==================================================
def generar_archivos_por_ruta(archivo_general, carpeta_salida, fecha_str):
    """
    Genera un archivo Excel por cada ruta,
    conservando únicamente su hoja correspondiente.
    """

    for ruta in ORDEN_RUTAS:
        nombre = f"Antigüedad {ruta} al {fecha_str}.xlsx"
        destino = os.path.join(carpeta_salida, nombre)

        shutil.copy(archivo_general, destino)
        wb = load_workbook(destino)

        for hoja in wb.sheetnames:
            if hoja != str(ruta):
                wb.remove(wb[hoja])

        wb.save(destino)