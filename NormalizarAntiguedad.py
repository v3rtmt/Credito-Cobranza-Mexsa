from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import numbers
from datetime import datetime
import re

def normalizar_antiguedad(path_antiguedad):
    """
    Normaliza un archivo de antigüedad de saldos en Excel y devuelve
    una lista de facturas con:
    - número de factura
    - prefijo
    - cliente
    - fecha
    - total
    """

    # =========================
    # Cargar archivo y hoja
    # =========================
    wb = load_workbook(path_antiguedad)
    hoja = wb.active

    # ==================================================
    # 1. Desunir celdas del encabezado (primeras filas)
    # ==================================================
    for rango in list(hoja.merged_cells):
        min_col, min_row, max_col, max_row = range_boundaries(str(rango))
        if min_row <= 9:
            hoja.unmerge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col
            )

    # Eliminar filas de encabezado innecesarias
    hoja.delete_rows(1, 11)

    # ==================================================
    # 2. Función auxiliar para desunir columnas específicas
    # ==================================================
    def desunir_hasta_columna(col_limite):
        rangos_a_desunir = [
            r for r in hoja.merged_cells.ranges if r.min_col <= col_limite
        ]

        for rango in rangos_a_desunir:
            try:
                hoja.unmerge_cells(str(rango))
            except Exception:
                pass

    desunir_hasta_columna(3)
    desunir_hasta_columna(5)

    # ==================================================
    # 3. Limpieza de columnas innecesarias
    # ==================================================
    hoja.delete_cols(5)
    hoja.delete_cols(3)
    hoja.delete_cols(1)

    # ==================================================
    # 4. Rellenar clientes vacíos (heredan el anterior)
    # ==================================================
    for fila in range(1, hoja.max_row + 1):
        if not hoja[f'B{fila}'].value:
            hoja[f'B{fila}'].value = hoja[f'B{fila - 1}'].value

    # ==================================================
    # 5. Eliminar filas que no tengan fecha válida
    # ==================================================
    for fila in range(hoja.max_row, 0, -1):
        fecha = hoja[f'D{fila}'].value
        if not re.match(r'^\d{2}/\d{2}/\d{4}$', str(fecha)):
            hoja.delete_rows(fila)

    # ==================================================
    # 6. Calcular el total sumando columnas de antigüedad
    # ==================================================
    hoja.insert_cols(5)  # Columna E para el total

    for fila in range(1, hoja.max_row + 1):
        if hoja[f'F{fila}'].value in (None, ''):
            break

        total = 0
        for col in range(6, 14):
            valor = hoja.cell(row=fila, column=col).value
            if isinstance(valor, (int, float)):
                total += valor

        hoja[f'E{fila}'].value = total
        hoja[f'E{fila}'].number_format = numbers.FORMAT_NUMBER

    # Eliminar columnas de antigüedad ya sumadas
    hoja.delete_cols(6, 12)

    # ==================================================
    # 7. Construcción de la lista de facturas
    # ==================================================
    facturas = []

    for fila in range(1, hoja.max_row + 1):
        factura_raw = hoja[f'A{fila}'].value

        if not factura_raw:
            continue

        if not str(factura_raw).startswith(('FACI', 'FACE')):
            continue

        facturas.append({
            'factura_raw': factura_raw,
            'factura': str(factura_raw).split('/')[-1],
            'prefijo': str(factura_raw).split('/')[1],
            'cliente_raw': hoja[f'B{fila}'].value,
            'fecha': datetime.strptime(hoja[f'C{fila}'].value, '%d/%m/%Y'),
            'total': hoja[f'E{fila}'].value
        })

    # ==================================================
    # 8. Ordenar por fecha (más reciente primero)
    # ==================================================
    facturas.sort(key=lambda x: x['fecha'], reverse=True)

    return facturas